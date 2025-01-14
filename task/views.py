import os
import shutil
import json
import random
import zipfile
import chardet
import string
import PIL.Image
import magic
import itertools
import openpyxl
from django.http import HttpRequest
from utils.utils_request import BAD_METHOD, request_failed, request_success, return_field
from utils.utils_require import MAX_TOKEN_LENGTH, MAX_TASKNAME_LENGTH, CheckRequire, require
from utils.utils_require import MAX_TEMPLATE_OBJECT_LIST_LENGTH,MAX_TEMPLATE_OBJECT_ROW_LENGTH,SUPPORTED_CONTENTS,MAX_TEMPLATENAME_LENGTH
from utils.utils_require import MAX_TAGS_COUNT,MAX_TAG_LENGTH,MAX_TEMPLATE_DESCRIPTION_LENGTH
from utils.utils_require import SUPPORTED_CONTENTS_WITH_DEFAULT,SUPPORTED_CONTENTS_WITHOUT_INPUTFILE,SUPPORTED_CONTENTS_WITHOUT_OUTPUTFILE
from utils.utils_require import MAX_UPLOAD_ZIP_SIZE,MAX_TEXT_CONTENT_LENGTH,MAX_TASK_LIST_COUNT
from utils.utils_time import get_timestamp
from user.models import User,AdminMessage
from task.models import TaskPackage, Task, TaskTemplate, TaskAnswer, Image, Tags
from user.userauth import userAuth
from django.core.paginator import Paginator,EmptyPage
from django.conf import settings
from task.tasks import delete_file_async
from utils.utils_user import maxAllValidateCount,sampleCount,minAcceptRate,defaultCredit,scoreDifference,maxPackagesAcceptForLowCredit
from utils.utils_user import proxyDifference
from user.views import send_message_to_user
from django.core.files import File
from django.core.files.images import ImageFile
from math import ceil,floor

# Create your views here.
@CheckRequire
def PackageInfo(req: HttpRequest, id):
    if req.method == "GET":
        body = req.GET.dict()

        # check params
        index = require({"id": id}, "id", "int", err_msg="Bad param [id]", err_code=-1)
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]", err_code=-2)
        package = TaskPackage.objects.filter(id=index).first()
        user = User.objects.filter(id=sender_id).first()

        # check auth
        if not package:
            return request_failed(1, "Package not found", status_code=404)
        if not user:
            return request_failed(1, "User not found", status_code=404)
        if len(sender_token) != MAX_TOKEN_LENGTH:
            return request_failed(-2, "Bad length of [token]", status_code=400)
        if user.token != sender_token:
            return request_failed(2, "Token not match", status_code=401)
        
        # check whether the user could see the package
        if not package.verified and user.id != package.creator.id and not user.adminPrivilege:
            return request_failed(2, "Permission denied", status_code=401)
        
        sender = User.objects.get(id=sender_id)
        sender.lastupdatedtime = get_timestamp()
        sender.save()

        return request_success(package.serialize())
    
    elif req.method == "DELETE":
        body = json.loads(req.body.decode())

        # check params
        index = require({"id": id}, "id", "int", err_msg="Bad param [id]", err_code=-1)
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]")
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]")
        package = TaskPackage.objects.filter(id=index).first()
        user = User.objects.filter(id=sender_id).first()

        # check auth
        if not package:
            return request_failed(1, "Package not found", status_code=404)
        if not user:
            return request_failed(1, "User not found", status_code=404)
        if len(sender_token) != MAX_TOKEN_LENGTH:
            return request_failed(-2, "Bad length of [token]", status_code=400)
        if user.token != sender_token:
            return request_failed(2, "Token not match", status_code=401)
        if not user.adminPrivilege and user != package.creator:
            return request_failed(2, "No privilege to delete the package", status_code=401)
        
        if package.distributed:
            return request_failed(-1,"Cannot delete a distributed package", status_code=400)

        package.delete()
        user.lastupdatedtime = get_timestamp()
        user.save()
        return request_success()
        
    elif req.method == 'PUT':
        body = json.loads(req.body.decode())
        # Basic User Auth
        authResult = userAuth(body)
        if authResult != None:
            return authResult
        # Param Type Check
        senderId = body['senderId']
        sender = User.objects.get(id=senderId)
        id = require({'id':id},'id','int','bad param [id]',err_code=-1)
        distribution = require(body,'distribution','string','bad param [distribution]')
        maxDistributedUser = require(body,'maxDistributedUser','int','bad param [maxDistributedUser]')
        if maxDistributedUser <= 0:
            return request_failed(-2,'bad param [maxDistributedUser]: param leq to 0')
        subtaskLimit = require(body,'subtaskLimit','float','bad param [subtaskLimit]')
        if subtaskLimit <=0:
            return request_failed(-2,'bad param [subtaskLimit] with subtasklimit leq to 0')
        score = require(body,'score','int','bad param [score]')
        if score <0:
            return request_failed(-2,'bad param [score] with score less than 0')
        deadline = require(body,'deadline','float','bad param [deadline]')
        taskName = require(body,'taskName','string','bad param [taskName]')
        if len(taskName) > MAX_TASKNAME_LENGTH or len(taskName) == 0:
            return request_failed(-2,'bad param [taskName] with invalid length')
        templateId = require(body,'templateId','int','bad param [templateId]')
        tags = require(body,'tags','list','bad param [tags]')
        if len(tags) > MAX_TAGS_COUNT:
            return request_failed(-2,'bad param [tags]:too much tags')
        for s in tags:
            str = require({'str':s},'str',err_msg='bad param [tags]:argument not string')
            if len(str) > MAX_TAG_LENGTH:
                return request_failed(-2,'bad param [tags]:too long tags')
        description = require(body,'description',err_msg='bad param [description]')
        if len(description) > MAX_TEMPLATE_DESCRIPTION_LENGTH:
            return request_failed(-2,'bad param [description]:too long description')
        # Param Validity Check
        targetPackage = TaskPackage.objects.filter(id=id)
        if not targetPackage.exists():
            return request_failed(1,'target package not found',404)
        targetPackage = targetPackage.first()
        if sender.adminPrivilege == False and sender != targetPackage.creator:
            return request_failed(2,'No permission',status_code=401)
        if 'proxied' in body:
            proxied = require(body,'proxied','bool','bad param [proxied]')
        else:
            proxied = False
        
        template = TaskTemplate.objects.filter(id=templateId)
        if not template.exists():
            return request_failed(1,'template not found',404)
        template = template.first()
        if not template.verified:
            return request_failed(-3, "The template has not been verified and cannot be used")
        if targetPackage.distributed:
            return request_failed(-2,'Cannot modify a distributed package!')
        targetPackage.tagsField.clear()
        # Modification
        targetPackage.distribution=distribution
        targetPackage.maxDistributedUser=maxDistributedUser
        targetPackage.subtaskLimit=subtaskLimit
        targetPackage.score=score
        targetPackage.deadline=deadline
        targetPackage.taskName=taskName
        targetPackage.template=template
        targetPackage.tags=tags
        targetPackage.description=description
        targetPackage.proxied=proxied
        for i in range(1,len(tags)+1):
            for subset in itertools.combinations(tags,i):
                tagset="+".join(sorted(subset))
                if not Tags.objects.filter(name=tagset).exists():
                    newtag=Tags(name=tagset)
                    newtag.save()
                targetPackage.tagsField.add(Tags.objects.get(name=tagset))
        targetPackage.save()
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        return request_success({})

    else:
        return BAD_METHOD

def check_for_create_data(body):
    sender_id = require(body, "senderId", "int", err_msg = "Missing or error type of [senderId]")
    sender_token = require(body, "senderToken", "string", err_msg = "Missing or error type of [senderToken]")
    assert len(sender_token) == MAX_TOKEN_LENGTH, "Bad length of [token]"

    distribution = require(body, "distribution", "int", err_msg = "Missing or error type of [distribution]")

    maxDistributedUser = require(body, "maxDistributedUser","int","bad param [maxDistributedUser]")

    assert maxDistributedUser > 0, "bad param [maxDistributedUser]: param leq to 0"

    subtask_limit = require(body, "subtaskLimit", "float", err_msg = "Missing or error type of [subtaskLimit]")
    assert subtask_limit > 0, "[subtaskLimit] should be positive"

    score = require(body, "score", "int", err_msg = "Missing or error type of [score]")
    assert score >= 0, "[score] should be positive"

    deadline = require(body, "deadline", "float", err_msg = "Missing or error type of [deadline]")
    assert deadline > 0, "[deadline] should be positive"

    task_name = require(body, "taskName", "string", err_msg = "Missing or error type of [taskName]")
    assert 0 < len(task_name) <= MAX_TASKNAME_LENGTH, "Bad length of [taskName]"

    template_id = require(body, "templateId", "int", err_msg = "Missing or error type of [templateId]")

    tags = require(body,'tags','list','bad param [tags]')
    assert len(tags) <= MAX_TAGS_COUNT,'bad param [tags]:too many tags'
    for s in tags:
        str = require({'str':s},'str',err_msg='bad param [tags]:argument not string')
        assert len(str) <= MAX_TAG_LENGTH,'bad param [tags]:too long tags'
    description = require(body,'description',err_msg='bad param [description]')
    assert len(description) <= MAX_TEMPLATE_DESCRIPTION_LENGTH,'bad param [description]:too long description'
    return sender_id, sender_token, distribution, maxDistributedUser, subtask_limit, score, deadline, task_name, template_id, tags, description

@CheckRequire
def createPackage(req: HttpRequest):
    if req.method == "POST":
        body = json.loads(req.body.decode("utf-8"))
        sender_id, sender_token, distribution, maxDistributedUser, subtask_limit, score, deadline, task_name, template_id, tags, description = check_for_create_data(body)

        change_user = User.objects.filter(id = sender_id).first()
        if not change_user:
            return request_failed(1, "sender does not exist", status_code = 404)

        if change_user.token != sender_token:
            return request_failed(2, "No permissions", status_code = 401)
        if change_user.adminPrivilege == False and change_user.uploadPrivilege == False:
            return request_failed(2, "No permissions", status_code = 401)

        template = TaskTemplate.objects.filter(id = template_id).first()
        if not template:
            return request_failed(1, "Template does not exist", status_code = 404)
        if not template.verified:
            return request_failed(-3, "The template has not been verified and cannot be used")
        if 'proxied' in body:
            proxied = require(body,'proxied','bool','bad param [proxied]')
        else:
            proxied = False

        task_package = TaskPackage(
            creator = change_user, 
            template = template, 
            distribution = distribution, 
            maxDistributedUser = maxDistributedUser,
            subtaskLimit = subtask_limit, 
            score = score, 
            deadline = deadline, 
            taskName = task_name,
            tags = tags,
            description = description,
            createTime = get_timestamp(),
            proxied = proxied)
        task_package.save()
        for i in range(1,len(tags)+1):
            for subset in itertools.combinations(tags,i):
                tagset="+".join(sorted(subset))
                if not Tags.objects.filter(name=tagset).exists():
                    newtag=Tags(name=tagset)
                    newtag.save()
                task_package.tagsField.add(Tags.objects.get(name=tagset))
        return request_success({'taskPackageId': task_package.id})

    else:
        return BAD_METHOD

@CheckRequire
def task_template_create_and_info(req: HttpRequest):
    if req.method=='POST':
        body = json.loads(req.body.decode())
        authres = userAuth(body)
        if authres != None:
            return authres
        sender = User.objects.get(id=body['senderId'])
        if not sender.adminPrivilege and not sender.uploadPrivilege:
            return request_failed(2,'no permission to create template',401)

        objectList = require(body,'objectList','list','bad param [objectList]')
        if len(objectList) > MAX_TEMPLATE_OBJECT_LIST_LENGTH or len(objectList) <= 0:
            return request_failed(-2,'bad param [objectList] with too many arguments')
        
        for c in objectList:
            if not isinstance(c,dict):
                return request_failed(-2,'bad param [objectList] with an object not dict')
            type = require(c,'type',err_msg='bad param [type] in objectList')
            if type not in SUPPORTED_CONTENTS:
                return request_failed(-2,'bad param [type] in objectList : type not supported')
            count = require(c,'count','int',err_msg='bad param [count] in objectList')
            if count <= 0 or count >MAX_TEMPLATE_OBJECT_ROW_LENGTH:
                return request_failed(-2,'bad param [count] in objectList : count length invalid')
            if 'setting' in c:
                settingContent = c['setting']
                if 'allowedMethod' in settingContent:
                    if type != 'imageinput':
                        return request_failed(-2,f'bad param [setting] in objectList: redundant "alloweMethod" param for element type {type}')
                    require(settingContent,'allowedMethod','list','bad param [setting] in objectList: invalid type of allowedMethod')
                    for elem in settingContent['allowedMethod']:
                        require({'elem':elem},'elem','int','bad param [setting] in objectList: invalid type of allowedMethod element')
                if 'default' in settingContent:
                    if not type in SUPPORTED_CONTENTS_WITH_DEFAULT:
                        return request_failed(-2,f"bad param [setting] in objectList: {type} does not support default value settings.")
                    require(settingContent,'default','list','bad param [setting] in objectList: "default" is not an array')
                    if len(settingContent['default'])!=count:
                        return request_failed(-2,'bad param [setting] in objectList: length of key "default" is not correct.')
                    for s in settingContent['default']:
                        require({'s':s},'s',err_msg='bad param [setting] in objectList: element in "default" is not string.')
                        if len(s)>MAX_TEXT_CONTENT_LENGTH:
                            return request_failed(-2,'bad param [setting] in objectList: element in "default" is too long')

            else:
                c['setting']={}
        templateName = require(body,'templateName',err_msg='bad param [templateName]')
        if len(templateName) <=0 or len(templateName)>MAX_TEMPLATENAME_LENGTH:
            return request_failed(-2,'bad param [templateName] with invalid length')
        description = require(body,"description",err_msg="bad param [description]")
        if len(description) > MAX_TEMPLATE_DESCRIPTION_LENGTH:
            return request_failed(-2,'bad param [description]: too long description')

        template = TaskTemplate(objectList=objectList,creator=sender,verified=False,templateName=templateName,description=description)
        template.save()
        message = AdminMessage(targetId=template.id,type="template",sender=sender)
        message.save()
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        return request_success({'templateId':template.id})
    elif req.method == 'GET':
        body = req.GET.dict()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        pageId = require(body,'pageId','int','bad param [pageId]')
        if pageId <= 0:
            return request_failed(-2,'bad param [pageId] with pageId leq 0')
        count = require(body,'count','int','bad param [count]')
        if count <= 0:
            return request_failed(-2,'bad param [count] with count leq 0')
        onlyVerified = require(body,'onlyVerified','int','bad param [onlyVerified]')
        if onlyVerified != 0 and onlyVerified != 1:
            return request_failed(-2,'bad param [onlyVerified] with not bool')
        templates = TaskTemplate.objects
        if onlyVerified:
            templates = templates.filter(verified = True)
        templatePages = Paginator(templates.all(),count)
        try:
            page = templatePages.page(pageId)
        except EmptyPage:
            page = templatePages.page(templatePages.num_pages)
        
        sender = User.objects.get(id=body['senderId'])
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        return request_success({'pageCount':templatePages.num_pages,'templateList':[
            return_field(t.serialize(), ["id", "creator", "verified", "templateName", "description"]) for t in page.object_list
        ]})
    else:
        return BAD_METHOD

@CheckRequire
def task_template_info(req: HttpRequest, id: any):
    if req.method == "GET":
        body = req.GET.dict()

        # check params
        index = require({"id": id}, "id", "int", err_msg="Bad param [id]", err_code=-1)
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]", err_code=-2)
        template = TaskTemplate.objects.filter(id=index).first()
        user = User.objects.filter(id=sender_id).first()

        if not template:
            return request_failed(1, "Package not found", status_code=404)
        if not user:
            return request_failed(1, "User not found", status_code=404)
        if len(sender_token) != MAX_TOKEN_LENGTH:
            return request_failed(-2, "Bad length of [token]", status_code=400)
        if user.token != sender_token:
            return request_failed(2, "Token not match", status_code=401)
        
        sender = User.objects.get(id=sender_id)
        if not template.verified and not sender.adminPrivilege and sender != template.creator:
            return request_failed(5,"The template is not verified",401)
        sender.lastupdatedtime = get_timestamp()
        sender.save()

        return request_success(template.serialize())
        
    else:
        return BAD_METHOD

@CheckRequire
def distribute_taskpackage(req: HttpRequest, id: any):
    if req.method == "POST":
        body = json.loads(req.body.decode())
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        packageId = require({"id":id},"id","int","Bad param [id]")
        package = TaskPackage.objects.filter(id=packageId)
        if not package.exists():
            return request_failed(1,"Task Package Not Found",404)
        package=package.first()
        if not package.uploaded:
            return request_failed(-1, "Task Package Not Uploaded", 400)
        if package.distributed:
            return request_failed(-1, "Task Package Already Distributed", 400)
        sender = User.objects.get(id=body['senderId'])
        if sender != package.creator:
            return request_failed(2,"No Privilege to distribute the package",401)
        if not package.proxied:
            score_needed = ceil(package.maxDistributedUser*package.score*scoreDifference)
            if sender.score < score_needed:
                return request_failed(-2, f"You don't have enough score to distribute. You need {score_needed}.")
            sender.score -= score_needed
            sender.experience += score_needed

            sender.lastupdatedtime = get_timestamp()
            sender.save()
            message = AdminMessage(targetId = package.id,type="package",sender=sender)
            message.save()
            package.availbleCount=package.maxDistributedUser
            package.distributed=True
            package.save()
            return request_success()
        else:
            score_needed = ceil(package.maxDistributedUser*package.score*(scoreDifference+proxyDifference))
            if sender.score < score_needed:
                return request_failed(-2, f"You don't have enough score to distribute. You need {score_needed}.")
            sender.score -= score_needed
            sender.experience += score_needed
            message = AdminMessage(targetId = package.id,type="package",sender=sender)
            message.save()
            sender.lastupdatedtime = get_timestamp()
            sender.save()
            package.availbleCount=1
            package.distributed=True
            package.save()
            return request_success()
    else:
        return BAD_METHOD

def handle_user_timeout(user: any):
    answers = TaskAnswer.objects.filter(submitter=user,finished=False)
    answers.delete()
    send_message_to_user(user.id,f"您未在规定的时限内完成编号为{user.currentTaskPackage.id}的任务，已扣除信用分。下次请抓紧时间。")
    if user.currentTaskPackage.proxied:
        user.currentTaskPackage.intermediary.credit-=1
        user.currentTaskPackage.intermediary.save()
        send_message_to_user(user.currentTaskPackage.intermediary.id,f"编号为{user.id}，用户名为{user.userName}的用户未在规定时限内完成任务，已扣除您的信用分。")
    user.refresh_from_db()
    user.currentTaskPackage.availbleCount+=1
    user.currentTaskPackage.save()
    user.currentTaskPackage=None
    user.deadline=1e15
    user.credit-=5
    user.save()

@CheckRequire
def get_taskpackage_todo(req: HttpRequest):
    if req.method == "GET":
        body = req.GET.dict()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        senderId = body['senderId']
        sender = User.objects.get(id=senderId)
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        # Handle the timeout users
        timeoutUsers = User.objects.filter(deadline__lt=get_timestamp())
        for user in timeoutUsers.all():
            handle_user_timeout(user)
        if 'mode' in body:
            mode=require(body,'mode',err_msg='bad param [mode]')
            if not mode in {'serial','preference','score','problemCount'}:
                return request_failed(-2,"bad param [mode]: mode not supported")
            if mode != 'serial' and sender.vipExpireTime<get_timestamp():
                return request_failed(2,f"You're not a vip user and cannot use mode {mode}.",401)
        else:
            mode='serial'
        if sender.currentTaskPackage != None:
            return request_success({"package":sender.currentTaskPackage.serialize(),"isAnAcceptedTask":True})
        elif sender.labelPrivilege == False:
            return request_success({"package":None,"isAnAcceptedTask":False})
        else:
            if mode=='preference':
                for i in range(len(sender.preferTags),0,-1):
                    for subset in itertools.combinations(sender.preferTags,i):
                        tagstr="+".join(sorted(subset))
                        if Tags.objects.filter(name=tagstr).exists():
                            packages = Tags.objects.get(name=tagstr).packagesWithTag.exclude(rejectedUser__in=User.objects.filter(id=senderId)).filter(verified=True,availbleCount__gt=0,proxied=False)
                            if packages.exists():
                                return request_success({"package":packages.first().serialize(),"isAnAcceptedTask":False})
                packages = TaskPackage.objects.exclude(rejectedUser__in=User.objects.filter(id=senderId)).filter(verified=True,availbleCount__gt=0,proxied=False)
            elif mode=='serial':
                packages = TaskPackage.objects.exclude(rejectedUser__in=User.objects.filter(id=senderId)).filter(verified=True,availbleCount__gt=0,proxied=False)
            elif mode=='score':
                packages = TaskPackage.objects.exclude(rejectedUser__in=User.objects.filter(id=senderId))\
                .filter(verified=True,availbleCount__gt=0,proxied=False)\
                .order_by('-score')
            elif mode=='problemCount':
                packages = packages = TaskPackage.objects.exclude(rejectedUser__in=User.objects.filter(id=senderId))\
                .filter(verified=True,availbleCount__gt=0,proxied=False)\
                .order_by('taskCount')
            else:
                return request_failed(1,'We have an unexpected error. Contact admin about this.',500)
            if packages.exists():
                return request_success({"package":packages.first().serialize(),"isAnAcceptedTask":False})
            else:
                return request_success({"package":None,"isAnAcceptedTask":False})
    else:
        return BAD_METHOD

@CheckRequire
def get_task_package_todo_agent(req: HttpRequest):
    if req.method=="GET":
        body = req.GET.dict()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        senderId = body['senderId']
        sender = User.objects.get(id=senderId)
        if not sender.mediationPrivilege:
            return request_failed(2,"No mediation privilege",401)
        if 'mode' in body:
            mode=require(body,'mode',err_msg='bad param [mode]')
            if not mode in {'serial','score'}:
                return request_failed(-2,"bad param [mode]: mode not supported")
            if mode != 'serial' and sender.vipExpireTime<get_timestamp():
                return request_failed(2,f"You're not a vip user and cannot use mode {mode}.",401)
        else:
            mode='serial'
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        packages = TaskPackage.objects.exclude(rejectedUser__in=User.objects.filter(id=senderId)).filter(verified=True,availbleCount__gt=0,proxied=True)
        if mode == 'score':
            packages=packages.order_by('-score')
        if packages.exists():
            return request_success({"package":packages.first().serialize()})
        else:
            return request_success({"package":None})
    else:
        return BAD_METHOD

@CheckRequire
def accept_taskpackage(req: HttpRequest,id: any):
    if req.method == "POST":
        id2 = require({"id": id}, "id", "int", err_msg = "Bad param [id]", err_code = -1)
        body = json.loads(req.body.decode())
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        sender = User.objects.get(id=body['senderId'])
        if not sender:
            return request_failed(1,"Sender not found",404)
        package = TaskPackage.objects.filter(id=id2)
        if not package.exists():
            return request_failed(1,"Package not found",404)
        package=package.first()
        accept = require(body,'accept','bool','bad param [accept]')

        
        if accept:
            if not sender.labelPrivilege:
                return request_failed(2,"No privilege to label tasks",401)
            if sender.currentTaskPackage != None:
                return request_failed(-2,"User has incompleted taskpackage. Please finish it first.",400)
            if package.availbleCount <= 0 or package.distributed == False or package.verified == False:
                return request_failed(-1,"Package not availble",400)
            if package.proxied:
                return request_failed(-2,"The package is a proxied package and cannot be accepted by normal user")
            
            newAcceptTime = []
            curtime=get_timestamp()
            for t in sender.packageAcceptTime:
                if curtime-t<24*60*60:
                    newAcceptTime.append(t)
            sender.packageAcceptTime=newAcceptTime
            if sender.credit <= 75 and len(sender.packageAcceptTime)>=maxPackagesAcceptForLowCredit:
                return request_failed(-3,"You accept too much packages in last 24 hours.")
            
            package.rejectedUser.add(sender)
            sender.currentTaskPackage=package
            sender.deadline=get_timestamp()+package.deadline
            package.availbleCount-=1
            sender.packageAcceptTime.append(get_timestamp())
        else:
            package.rejectedUser.add(sender)
        sender.lastupdatedtime=get_timestamp()
        sender.save()
        package.save()
        return request_success()
    else:
        return BAD_METHOD
    
@CheckRequire
def accept_taskpackage_agent(req: HttpRequest,id: any):
    if req.method == "POST":
        id2 = require({"id": id}, "id", "int", err_msg = "Bad param [id]", err_code = -1)
        body = json.loads(req.body.decode())
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        sender = User.objects.get(id=body['senderId'])
        if not sender:
            return request_failed(1,"Sender not found",404)
        package = TaskPackage.objects.filter(id=id2)
        if not package.exists():
            return request_failed(1,"Package not found",404)
        package=package.first()
        
        accept = require(body,'accept','bool','bad param [accept]')
        if not sender.mediationPrivilege:
            return request_failed(2,"No mediation privilege",401)
        if package.availbleCount <= 0 or package.uploaded == False:
            return request_failed(-1,"Package not availble",400)
        if not package.proxied:
            return request_failed(-2,"The package is not a proxied package and cannot be accpeted by agents")
        
        newAcceptTime = []
        curtime=get_timestamp()
        for t in sender.packageAcceptTime:
            if curtime-t<24*60*60:
                newAcceptTime.append(t)
        sender.packageAcceptTime=newAcceptTime
        sender.save()
        if sender.credit <= 75 and len(sender.packageAcceptTime)>=maxPackagesAcceptForLowCredit:
            return request_failed(-3,"You accept too much packages in last 24 hours.")

        if accept:
            targetUser = require(body,'targetUser','list','bad param [targetUser]')
            if len(targetUser)<package.maxDistributedUser:
                return request_failed(-2,"bad param [targetUser]:user count less than required")
            if len(targetUser)>package.maxDistributedUser:
                return request_failed(-2,"bad param [targetUser]:user count more than required")
            for userId in targetUser:
                userId = require({"userId":userId},"userId","int","bad param [targetUser]:cannot identify id")
                user = User.objects.filter(id=userId).first()
                if user == None:
                    return request_failed(-3,"bad param [targetUser]:user not found")
                if not user.labelPrivilege:
                    return request_failed(-3,"bad param [targetUser]:user has no privilege to label task")
                if user.currentTaskPackage != None:
                    return request_failed(-3,"bad param [targetUser]:user has incompleted package")
                if user.credit <= 75:
                    return request_failed(-3,"bad param [targetUser]:user has too low credit")
            package.rejectedUser.add(sender)
            for userId in targetUser:
                user = User.objects.get(id=userId)
                user.currentTaskPackage=package
                user.deadline=get_timestamp()+package.deadline
                user.save()
                send_message_to_user(user.id,f"编号为{sender.id}的中介向您派发了任务，请您及时完成。如有问题请向管理员举报。")
            package.availbleCount-=1
            package.intermediary=sender
            sender.refresh_from_db()
            sender.packageAcceptTime.append(get_timestamp())
        else:
            package.rejectedUser.add(sender)
            sender.refresh_from_db()
        sender.lastupdatedtime=get_timestamp()
        sender.save()
        package.save()
        return request_success()
    else:
        return BAD_METHOD

@CheckRequire
def get_next_task(req: HttpRequest):
    if req.method == "GET":
        body = req.GET.dict()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        sender = User.objects.get(id=body['senderId'])
        if sender.deadline < get_timestamp():
            handle_user_timeout(sender)
            return request_failed(-1,"Your current taskpackage has passed deadline. Get another task.",400)
        if sender.labelPrivilege == False:
            return request_failed(2,"No privilege to label task",401)
        if sender.currentTaskPackage == None:
            return request_failed(-1,"You haven't get a task or task has passed deadline. Get another task.",400)
        
        ongoingTask = TaskAnswer.objects.filter(submitter=sender,finished=False)
        unfinishedTasks = Task.objects.filter(packageBelonging=sender.currentTaskPackage).exclude(answers=sender)
        if ongoingTask.exists():
            return request_success({"taskId":ongoingTask.first().task.id,"count":unfinishedTasks.count()})
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        if not unfinishedTasks.exists():
            sender.currentTaskPackage.completedUser.add(sender)
            sender.currentTaskPackage=None
            sender.deadline=1e15
            sender.save()
            return request_success({"taskId":None,"count":0})
        task = unfinishedTasks.first()
        return request_success({"taskId":task.id,"count":unfinishedTasks.count()})
    else:
        return BAD_METHOD

@CheckRequire
def upload_data(req: HttpRequest, id: any):
    if req.method == "POST":
        # TODO: remove the redundant images after upload
        package_id = require({"id": id}, "id", "int", err_msg = "Bad param [id]", err_code = -1)
        package = TaskPackage.objects.filter(id = package_id).first()
        if not package:
            return request_failed(1, "TaskPackage not found", status_code = 404)
        
        sender_id = req.POST.get("senderId")
        sender_token = req.POST.get("senderToken")
        sender = User.objects.filter(id = sender_id).first()

        if not sender:
            return request_failed(1, "Sender not found", status_code = 404)
        if sender.token != sender_token:
            return request_failed(2, "No permissions", status_code = 401)
        if sender.uploadPrivilege == False and sender.adminPrivilege == False:
            return request_failed(2, "No permissions", status_code = 401)
        if sender != package.creator and not sender.adminPrivilege:
            return request_failed(2, "No permissions", status_code = 401)

        _file = req.FILES.get("uploadFile", None)
        if not _file:
            return request_failed(-2, "No files for upload", status_code = 400)
        
        folder_path = os.path.join(settings.MEDIA_ROOT, id)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        else:
            shutil.rmtree(folder_path)
            os.mkdir(folder_path)
        
        def request_failed_with_file_delete(code:int,message:str,status_code = 400):
            shutil.rmtree(folder_path)
            return request_failed(code,message,status_code)
        
        file_path = os.path.join(folder_path, _file.name)

        upload_file = open(file_path, "wb")
        for chunk in _file.chunks():
            upload_file.write(chunk)
        upload_file.close()

        if not zipfile.is_zipfile(file_path):
            return request_failed_with_file_delete(-3, "Upload file is not a zip", status_code = 400)

        with zipfile.ZipFile(file_path) as zfile:
            totFileSize=sum(e.file_size for e in zfile.infolist())
            if totFileSize > MAX_UPLOAD_ZIP_SIZE:
                return request_failed_with_file_delete(-3, "Upload file is too large after extraction")
            zfile.extractall(folder_path)

        idx = 1
        file_list = os.listdir(folder_path)
        while str(idx) in file_list:
            idx = idx + 1
        if idx + int("answer" in file_list) != len(file_list):
            return request_failed_with_file_delete(-3, "Upload file for task is wrong form", status_code = 400)

        template = package.template

        for i in range(idx - 1):
            row = 0
            sum_count = 0
            path = os.path.join(folder_path, str(i + 1))

            answer_file = os.path.join(folder_path, "answer", str(i + 1))
            
            for uni in template.objectList:
                row = row + 1
                uni_type = uni["type"]
                uni_count = uni["count"]
                if not uni_type in SUPPORTED_CONTENTS_WITHOUT_INPUTFILE:
                    sum_count += uni_count

                for line in range(uni_count):
                    if os.path.exists(answer_file):
                        if uni_type == "textinput":
                            if not os.path.exists(os.path.join(answer_file, str(row) + "_" + str(line + 1) + ".txt")):
                                return request_failed_with_file_delete(-3, "Upload answer_file in %d for textinput does not exist"%(i + 1), status_code = 400)

                        elif uni_type == "singlechoice":
                            if not os.path.exists(os.path.join(answer_file, str(row) + "_" + str(line + 1) + ".txt")):
                                return request_failed_with_file_delete(-3, "Upload answer_file in %d for singlechoice does not exist"%(i + 1), status_code = 400)

                        elif uni_type == "multiplechoice":
                            if not os.path.exists(os.path.join(answer_file, str(row) + "_" + str(line + 1) + ".txt")):
                                return request_failed_with_file_delete(-3, "Upload answer_file in %d for muliplechoice does not exist"%(i + 1), status_code = 400)

                    if uni_type == "text":
                        text_exists = os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt"))
                        if not text_exists\
                            and not ('setting' in uni and 'default' in uni['setting']):
                            return request_failed_with_file_delete(-3, "Upload file in %d for text is wrong form"%(i + 1), status_code = 400)
                        if text_exists:
                            with open(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encodingRes = chardet.detect(file_string)
                                if encodingRes['confidence'] >= 0.8:
                                    encoding = chardet.detect(file_string)["encoding"]
                                else:
                                    encoding = 'utf-8'
                                try:
                                    file_string = file_string.decode(encoding)
                                except:
                                    return request_failed_with_file_delete(-2,f"We cannot decode the txt at {i}-{row}-{line}. Try to use utf-8 decoding and submit again.")
                                if len(file_string) > MAX_TEXT_CONTENT_LENGTH:
                                    return request_failed_with_file_delete(-2,"Too long text in %d"%(i + 1), status_code = 400)
                    
                    elif uni_type == "image" or uni_type == "imageinput":
                        if not os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")) and not os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".png")):
                            return request_failed_with_file_delete(-3, "Upload file in %d for "+uni_type+" is wrong form"%(i + 1), status_code = 400)
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")
                            try:
                                imageContent = PIL.Image.open(image_path)
                                imageContent.verify()
                            except:
                                return request_failed_with_file_delete(-3, "Upload image in %d for "+uni_type+" is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".png")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".png")
                            try:
                                imageContent = PIL.Image.open(image_path)
                                imageContent.verify()
                            except:
                                return request_failed_with_file_delete(-3, "Upload image in %d for "+uni_type+" is broken or in wrong format"%(i+1))
                            
                    elif uni_type == "video":
                        if not os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4")) and not os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi")):
                            return request_failed_with_file_delete(-3, "Upload file in %d for video is wrong form"%(i + 1), status_code = 400)
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4"), mime=True) != 'video/mp4':
                                return request_failed_with_file_delete(-3, "Upload video in %d is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi"), mime=True) != 'video/x-msvideo':
                                return request_failed_with_file_delete(-3, "Upload video in %d is broken or in wrong format"%(i+1))
                    
                    elif uni_type == "audio":
                        if not os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav")) and not os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3")):
                            return request_failed_with_file_delete(-3, "Upload file in %d for audio is wrong form"%(i + 1), status_code = 400)
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav"), mime=True) != 'audio/x-wav':
                                return request_failed_with_file_delete(-3, "Upload audio in %d is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3"), mime=True) != 'audio/mpeg':
                                return request_failed_with_file_delete(-3, "Upload audio in %d is broken or in wrong format"%(i+1))
                    
                    elif uni_type == "multimedia":
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt")):
                            with open(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encodingRes = chardet.detect(file_string)
                                if encodingRes['confidence'] >= 0.8:
                                    encoding = chardet.detect(file_string)["encoding"]
                                else:
                                    encoding = 'utf-8'
                                try:
                                    file_string = file_string.decode(encoding)
                                except:
                                    return request_failed_with_file_delete(-2,f"We cannot decode the txt at {i}-{row}-{line}. Try to use utf-8 decoding and submit again.")
                                if len(file_string) > MAX_TEXT_CONTENT_LENGTH:
                                    return request_failed_with_file_delete(-2,"Too long text in %d"%(i + 1), status_code = 400)
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")
                            try:
                                imageContent = PIL.Image.open(image_path)
                                imageContent.verify()
                            except:
                                return request_failed_with_file_delete(-3, "Upload image in %d for multimedia is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".png")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".png")
                            try:
                                imageContent = PIL.Image.open(image_path)
                                imageContent.verify()
                            except:
                                return request_failed_with_file_delete(-3, "Upload image in %d for multimedia is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4"), mime=True) != 'video/mp4':
                                return request_failed_with_file_delete(-3, "Upload video in %d is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi"), mime=True) != 'video/x-msvideo':
                                return request_failed_with_file_delete(-3, "Upload video in %d is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav"), mime=True) != 'audio/x-wav':
                                return request_failed_with_file_delete(-3, "Upload audio in %d is broken or in wrong format"%(i+1))
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3")):
                            if magic.from_file(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3"), mime=True) != 'audio/mpeg':
                                return request_failed_with_file_delete(-3, "Upload audio in %d is broken or in wrong format"%(i+1))
                        else:
                            return request_failed_with_file_delete(-3, "Upload file in %d for multimedia is wrong form"%(i + 1))
                            
                if os.path.exists(answer_file):
                    if uni_type == "singlechoice":
                        many = 0
                        for line in range(uni_count):
                            with open(os.path.join(answer_file, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encodingRes = chardet.detect(file_string)
                                if encodingRes['confidence'] >= 0.8:
                                    encoding = chardet.detect(file_string)["encoding"]
                                else:
                                    encoding = 'utf-8'
                                file_string = file_string.decode(encoding)

                                if file_string == "True":
                                    many = many + 1
                                elif file_string != "False":
                                    return request_failed_with_file_delete(-3, "Upload answer_file in %d for singlechoice has invalid words"%(i + 1), status_code = 400)
                        
                        if many != 1:
                            return request_failed_with_file_delete(-3, "Upload answer_file in %d for singlechoice has error in the number of options"%(i + 1), status_code = 400)

                    elif uni_type == "multiplechoice":
                        for line in range(uni_count):
                            with open(os.path.join(answer_file, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encodingRes = chardet.detect(file_string)
                                if encodingRes['confidence'] >= 0.8:
                                    encoding = chardet.detect(file_string)["encoding"]
                                else:
                                    encoding = 'utf-8'
                                file_string = file_string.decode(encoding)

                                if file_string != "True" and file_string != "False":
                                    return request_failed_with_file_delete(-3, "Upload answer_file in %d for singlechoice has invalid words"%(i + 1), status_code = 400)
            
            file_list = os.listdir(path)
            if len(file_list) > sum_count:
                return request_failed_with_file_delete(-3, "Upload file has useless file in %d"%(i + 1), status_code = 400)
        
        Task.objects.filter(packageBelonging = package).all().delete()
        
        for i in range(idx - 1):
            task = Task(packageBelonging = package, problemId = i)

            row = 0
            path = os.path.join(folder_path, str(i + 1))

            answer_file = os.path.join(folder_path, "answer/" + str(i + 1))
            if os.path.exists(answer_file):
                task.hasStandardAnswer = True

            for uni in template.objectList:
                row = row + 1
                uni_type = uni["type"]
                uni_count = uni["count"]

                ans_list = []

                for line in range(uni_count):
                    if os.path.exists(answer_file):
                        if uni_type == "textinput" or uni_type == "singlechoice" or uni_type == "multiplechoice":
                            with open(os.path.join(answer_file, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encoding = chardet.detect(file_string)["encoding"]
                                file_string = file_string.decode(encoding)
                                if uni_type == "singlechoice" or uni_type == "multiplechoice":
                                    ans_list.append({"data": file_string=='True'})
                                else:
                                    ans_list.append({"data": file_string})
            
                    if uni_type == "text":
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt")):
                            with open(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encodingRes = chardet.detect(file_string)
                                if encodingRes['confidence'] >= 0.8:
                                    encoding = chardet.detect(file_string)["encoding"]
                                else:
                                    encoding = 'utf-8'
                                try:
                                    file_string = file_string.decode(encoding)
                                except:
                                    return request_failed_with_file_delete(-2,f"We cannot decode the txt at {i}-{row}-{line}. Try to use utf-8 decoding and submit again.")
                        else:
                            file_string = uni['setting']['default'][line]
                        task.data.append({
                            "location": str(row) + "_" + str(line + 1),
                            "type": uni_type,
                            "text": file_string,
                        })
                    
                    elif uni_type == "image" or uni_type == "imageinput":
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")
                            image_name = "".join(random.choices(string.digits+string.ascii_letters,k=32))+".jpg"

                            with open(image_path, 'rb') as f:
                                django_file = File(f)
                                django_image_file = ImageFile(django_file)
                                _image = Image()
                                _image.image.save(image_name, django_image_file, save=True)

                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":uni_type,
                                "url": _image.image.url,
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".png")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".png")
                            image_name = "".join(random.choices(string.digits+string.ascii_letters,k=32))+".png"

                            with open(image_path, 'rb') as f:
                                django_file = File(f)
                                django_image_file = ImageFile(django_file)
                                _image = Image()
                                _image.image.save(image_name, django_image_file, save=True)

                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "url": _image.image.url,
                            })
                    
                    elif uni_type == "video":
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".mp4"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":uni_type,
                                "url": "/media/"+newFileName
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".avi"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":uni_type,
                                "url": "/media/"+newFileName
                            })
                    
                    elif uni_type == "audio":
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".mp3"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":uni_type,
                                "url": "/media/"+newFileName
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".wav"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":uni_type,
                                "url": "/media/"+newFileName
                            })
                    elif uni_type == "multimedia":
                        if os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt")):
                            with open(os.path.join(path, str(row) + "_" + str(line + 1) + ".txt"), 'rb') as f:
                                file_string = f.read()
                                encodingRes = chardet.detect(file_string)
                                if encodingRes['confidence'] >= 0.8:
                                    encoding = chardet.detect(file_string)["encoding"]
                                else:
                                    encoding = 'utf-8'
                                try:
                                    file_string = file_string.decode(encoding)
                                except:
                                    return request_failed_with_file_delete(-2,f"We cannot decode the txt at {i}-{row}-{line}. Try to use utf-8 decoding and submit again.")
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type": "text",
                                "text": file_string,
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".jpg")
                            image_name = "".join(random.choices(string.digits+string.ascii_letters,k=32))+".jpg"

                            with open(image_path, 'rb') as f:
                                django_file = File(f)
                                django_image_file = ImageFile(django_file)
                                _image = Image()
                                _image.image.save(image_name, django_image_file, save=True)

                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":"image",
                                "url": _image.image.url,
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".png")):
                            image_path = os.path.join(path, str(row) + "_" + str(line + 1) + ".png")
                            image_name = "".join(random.choices(string.digits+string.ascii_letters,k=32))+".png"

                            with open(image_path, 'rb') as f:
                                django_file = File(f)
                                django_image_file = ImageFile(django_file)
                                _image = Image()
                                _image.image.save(image_name, django_image_file, save=True)

                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":"image",
                                "url": _image.image.url,
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".mp4"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp4"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":"video",
                                "url": "/media/"+newFileName
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".avi"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".avi"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":"video",
                                "url": "/media/"+newFileName
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".mp3"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".mp3"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":"audio",
                                "url": "/media/"+newFileName
                            })
                        elif os.path.exists(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav")):
                            newFileName = "".join(random.choices(string.ascii_letters+string.digits,k=32))+".wav"
                            shutil.copyfile(os.path.join(path, str(row) + "_" + str(line + 1) + ".wav"),os.path.join(settings.MEDIA_ROOT,newFileName))
                            task.data.append({
                                "location": str(row) + "_" + str(line + 1),
                                "type":"audio",
                                "url": "/media/"+newFileName
                            })
                task.standardAnswer.append(ans_list)

            task.save()
        
        newZipPath = "media/" + ("".join(random.choices(string.digits + string.ascii_letters, k = 32)) + ".zip")
        os.rename(file_path, newZipPath)
        shutil.rmtree(folder_path)
        
        package.uploaded = True
        if package.uploadedFileUrl != "":
            os.remove(os.path.join(settings.BASE_DIR, package.uploadedFileUrl))
        package.uploadedFileUrl = newZipPath
        package.taskCount = idx - 1
        package.save()
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        return request_success()

    else:
        return BAD_METHOD

def read_file(file_name, chunk_size = 512):
    with open(file_name, "rb") as f:
        while True:
            c = f.read(chunk_size)
            if c:
                yield c
            else:
                break
    f.close()

@CheckRequire
def download_data(req: HttpRequest, id: any):
    if req.method == "GET":
        package_id = require({"id": id}, "id", "int", err_msg = "Bad param [id]", err_code = -1)
        package = TaskPackage.objects.filter(id = package_id).first()
        if not package:
            return request_failed(1, "TaskPackage not found", status_code = 404)
    
        body = req.GET.dict()
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]", err_code=-2)

        sender = User.objects.filter(id = sender_id).first()
        if not sender:
            return request_failed(1, "Sender not found", status_code = 404)
        if sender.token != sender_token:
            return request_failed(2, "No permissions", status_code = 401)
        if sender != package.creator and not sender.adminPrivilege:
            return request_failed(2, "No permissions", status_code = 401)

        return request_success({"filePath":package.uploadedFileUrl})

    else:
        return BAD_METHOD

@CheckRequire
def task_info(req: HttpRequest, id: any):
    if req.method == "GET":
        task_id = require({"id": id}, "id", "int", err_msg = "Bad param [id]", err_code = -1)
        task = Task.objects.filter(id = task_id).first()
        if not task:
            return request_failed(1, "Task not found", status_code = 404)

        body = req.GET.dict()
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]", err_code=-2)

        sender = User.objects.filter(id = sender_id).first()
        if not sender:
            return request_failed(1, "Sender not found", status_code = 404)
        if sender.token != sender_token:
            return request_failed(2, "No permissions", status_code = 401)
        if not sender.adminPrivilege and not sender.labelPrivilege:
            return request_failed(2, "No permissions", status_code = 401)
        if sender.adminPrivilege == False\
            and sender.currentTaskPackage != task.packageBelonging and sender != task.packageBelonging.creator:
            return request_failed(2, "No permissions", status_code = 401)

        if sender.currentTaskPackage == task.packageBelonging:
            if get_timestamp() > sender.deadline:
                return request_failed(2, "Deadline has passed", status_code = 401)

            task_answer = TaskAnswer.objects.filter(submitter = sender, finished = False).first()
            if task_answer and task_answer.task != task:
                return request_failed(-2, "Please complete the unfinished tasks", status_code = 400)

            task_answer = TaskAnswer.objects.filter(submitter = sender, task = task).first()
            if task_answer and task_answer.finished == True:
                return request_failed(-2, "You have finished the task", status_code = 400)

            if not task_answer:
                task_answer = TaskAnswer(submitter = sender, task = task, startTime = get_timestamp())
            else:
                task_answer.startTime = get_timestamp()
            task_answer.save()

        sender.lastupdatedtime = get_timestamp()
        sender.save()

        return request_success(task.serialize())

    else:
        return BAD_METHOD

@CheckRequire
def submit_answer(req: HttpRequest, id: any):
    if req.method == "PUT":
        body = json.loads(req.body.decode())
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        sender = User.objects.get(id=body['senderId'])
        if not sender.labelPrivilege:
            return request_failed(2,"No privilege to label",401)
        taskId = require({"id":id},"id","int","bad param [id]",-1)
        task=Task.objects.filter(id=taskId)
        if not task.exists():
            return request_failed(1,"Package not found",404)
        task=task.first()
        answer=TaskAnswer.objects.filter(submitter=sender,task=task)
        if not answer.exists():
            return request_failed(-2,"Task haven't started or package not accepted")
        answer=answer.first()
        if answer.finished:
            return request_failed(-2,"Answer already submitted")
        if answer.startTime+task.packageBelonging.subtaskLimit>get_timestamp():
            return request_failed(-2,"You finished the task too fast. Do it more carefully.")
        # validate the format of the answer
        answerPayload = require(body,"answers","list","bad param [answers]")
        template = task.packageBelonging.template
        if len(answerPayload) != len(template.objectList):
            return request_failed(-2,"format of answer doesn't match with template [0x1]")
        for i in range(0,len(answerPayload)):
            subPayload = require({"payload":answerPayload[i]},"payload","list","format of answer doesn't match with template [0x2]")
            unitType = template.objectList[i]['type']
            if unitType in SUPPORTED_CONTENTS_WITHOUT_OUTPUTFILE:
                if len(subPayload)!=0:
                    return request_failed(-2,"format of answer doesn't match with template [0x3]")
                continue
            if len(subPayload)!=template.objectList[i]['count']:
                return request_failed(-2,"format of answer doesn't match with template [0x3]")
            choiceCnt=0            
            for j in range(0,len(subPayload)):
                blockAnswer = require({"blockAnswer":subPayload[j]},"blockAnswer","dict","format of answer doesn't match with template [0x4]")

                # check the validity of the answer (blockanswer just has data now)
                if len(blockAnswer) != 1:
                    return request_failed(-2,f"format of answer doesn't match with template [0x5]:{blockAnswer}")

                if unitType == "textinput":
                    require(blockAnswer,"data","string","format of answer doesn't match with template [0x5]")
                elif unitType == "singlechoice":
                    require(blockAnswer,"data","bool","format of answer doesn't match with template [0x5]")
                    choiceCnt += bool(blockAnswer["data"])
                elif unitType == "multiplechoice":
                    require(blockAnswer,"data","bool","format of answer doesn't match with template [0x5]")
                # TODO: check the validity of imageinput
            if unitType == "singlechoice" and choiceCnt != 1:
                return request_failed(-2,f"format of answer doesn't match with template [0x6]:choice of cnt is {choiceCnt}")
        answer.finished=True
        answer.payload=answerPayload
        answer.save()
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        return request_success()
    else:
        return BAD_METHOD

@CheckRequire
def get_taskpackage_list(req: HttpRequest):
    if req.method == "GET":
        body = req.GET.dict()
        
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]", err_code=-2)
        page_id = require(body, "pageId", "int", err_msg="Missing or error type of [pageId]", err_code=-2)
        count = require(body, "count", "int", err_msg="Missing or error type of [count]", err_code=-2)

        if page_id <= 0:
            return request_failed(-2, "Bad param [pageId]", status_code = 400)
        if count <= 0:
            return request_failed(-2, "Bad param [count]", status_code = 400)
        packages = TaskPackage.objects.all().order_by("-id")
        pages = Paginator(packages, count)
        try:
            page = pages.page(page_id)
        except EmptyPage:
            page = pages.page(pages.num_pages)
        
        sender = User.objects.filter(id = sender_id).first()
        if not sender:
            return request_failed(1, "Sender not found", status_code = 404)
        if sender.token != sender_token or not sender.adminPrivilege:
            return request_failed(2, "No permissions", status_code = 401)
        sender.lastupdatedtime = get_timestamp()
        sender.save()

        return request_success({'packageList': [package.serialize() for package in page.object_list],'pageCount':pages.num_pages})
    else:
        return BAD_METHOD
    
@CheckRequire
def get_created_package(req: HttpRequest):
    if req.method == "GET":
        body = req.GET.dict()
        
        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender_token = require(body, "senderToken", "string", err_msg="Missing or error type of [senderToken]", err_code=-2)

        sender = User.objects.filter(id = sender_id).first()
        if not sender:
            return request_failed(1, "Sender not found", status_code = 404)
        if sender.token != sender_token:
            return request_failed(2, "No permissions", status_code = 401)
        sender.lastupdatedtime = get_timestamp()
        sender.save()

        not_uploaded = TaskPackage.objects.filter(creator = sender, uploaded = False).order_by("-id")
        not_distributed = TaskPackage.objects.filter(creator = sender, uploaded = True, distributed = False).order_by("-id")
        distributed = TaskPackage.objects.filter(creator = sender, distributed = True).order_by("-id")

        return request_success({'notUploaded': [package.serialize() for package in not_uploaded], 'notDistributed': [package.serialize() for package in not_distributed], 'Distributed': [package.serialize() for package in distributed]})
    else:
        return BAD_METHOD

def result_bonus_calculate(submitterId: int, packageId: int, accepted: bool):
    submitter = User.objects.get(id=submitterId)
    package = TaskPackage.objects.get(id=packageId)
    if accepted:
        submitter.experience+=package.score
        submitter.score+=package.score
        submitter.availbleScore+=package.score
        submitter.labelCount+=1
        if submitter.credit < defaultCredit:
            submitter.credit += 1
        submitter.save()
        package.completedCount+=1
        package.save()
        send_message_to_user(submitter.id,f"您完成的编号为{packageId}的任务已通过审核。")
        if package.completedCount == package.maxDistributedUser:
            send_message_to_user(package.creator.id,f"编号为{package.id}的任务已经完成全部审核。")
            if package.proxied:
                send_message_to_user(package.intermediary.id,f"编号为{package.id}的任务已经完成全部审核，您的提成已经发放到账户。")
                mscore = floor(package.verifiedUser.count()*package.score*proxyDifference)
                package.intermediary.experience += mscore
                package.intermediary.score += mscore
                package.intermediary.availbleScore += mscore
                if package.intermediary.credit < defaultCredit:
                    package.intermediary.credit += 1
                if package.verifiedUser.count()!=package.maxDistributedUser:
                    package.creator.score+=floor((package.maxDistributedUser-package.verifiedUser.count())*package.score*(proxyDifference+scoreDifference))
                    package.creator.save()
                    send_message_to_user(package.creator.id,f"编号为{package.id}的任务有{package.maxDistributedUser-package.verifiedUser.count()}个用户未能成功完成，退款已发放到账户。")
    else:
        submitter.credit -= 5
        submitter.save()
        send_message_to_user(submitter.id,f"您完成的编号为{packageId}的任务未通过审核。如果对结果有异议请自行联系需求方或向管理员举报。")
        if package.proxied:
            package.intermediary.credit-=1
            package.intermediary.save()
            send_message_to_user(package.intermediary.id,f"编号为{submitter.id}，用户名为{submitter.userName}的用户完成的任务未通过审核，已扣除您的信用分。")
            package.completedCount+=1
            package.save()

def calAcceptRate(sender: any):
    if sender.credit >= 90:
        return minAcceptRate
    if sender.credit <= 75:
        return 1
    return (minAcceptRate*(sender.credit-75)+(90-sender.credit))/15

@CheckRequire
def manual_validate(req: HttpRequest, id: any):
    if req.method == "GET":
        body = req.GET.dict()
        packageId = require({"id":id},"id","int","bad param [id]",-1)
        package = TaskPackage.objects.filter(id=packageId)
        if not package.exists():
            return request_failed(1,"Package not found",404)
        package = package.first()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        mode = require(body,"mode",err_msg="bad param [mode]")
        if mode != "all" and mode != "sampling":
            return request_failed(-2,"Unsupported vallidation mode",-1)
        submitterId = require(body,"submitterId","int","bad param [submitterId]")
        submitter = User.objects.filter(id=submitterId)
        if not submitter.exists():
            return request_failed(1,"Submitter not found",404)
        submitter = submitter.first()
        if not package.completedUser.filter(id=submitter.id).exists():
            return request_failed(-2,"The submitter is not in the validation list")

        sender = User.objects.get(id=body['senderId'])
        if sender != package.creator:
            return request_failed(2,"No permission to validate package",401)
        
        tasks=[]
        if mode == "all":
            if Task.objects.filter(packageBelonging=package).count()>maxAllValidateCount:
                return request_failed(-2,"Too much problems for a overall validation")
            tasks=[a.serialize() for a in TaskAnswer.objects.filter(submitter=submitter,task__packageBelonging=package).all()]
        elif mode == "sampling":
            answers = list(TaskAnswer.objects.filter(submitter=submitter,task__packageBelonging=package).all())
            tasks=[a.serialize() for a in random.sample(answers,k=min(sampleCount,len(answers)))]
        sender = User.objects.get(id=body['senderId'])
        sender.lastupdatedtime=get_timestamp()
        sender.save()
        return request_success({"tasks":tasks})
    elif req.method == "PUT":
        body = json.loads(req.body.decode())
        packageId = require({"id":id},"id","int","bad param [id]",-1)
        package = TaskPackage.objects.filter(id=packageId)
        if not package.exists():
            return request_failed(1,"Package not found",404)
        package = package.first()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        submitterId = require(body,"submitterId","int","bad param [submitterId]")
        submitter = User.objects.filter(id=submitterId)
        if not submitter.exists():
            return request_failed(1,"Submitter not found",404)
        submitter = submitter.first()
        if not package.completedUser.filter(id=submitter.id).exists():
            return request_failed(-2,"The submitter is not in the validation list")
        acceptRate = require(body,"acceptRate","float","bad param [acceptRate]")
        if acceptRate <0 or acceptRate > 1:
            return request_failed(-2,"bad param [acceptRate] with acceptRate larger than 1 or less than 0")
        sender = User.objects.get(id=body['senderId'])
        if sender != package.creator:
            return request_failed(2,"No permission to validate package",401)
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        package.completedUser.remove(submitter)
        if acceptRate >= calAcceptRate(submitter):
            package.verifiedUser.add(submitter)
            package.save()
            if package.verifiedUser.count() == package.maxDistributedUser and sender.credit < defaultCredit:
                sender.credit += 1
            result_bonus_calculate(submitterId,packageId,True)
            return request_success({"accepted":True})
        else:
            package.availbleCount += 1
            package.save()
            result_bonus_calculate(submitterId,packageId,False)
            return request_success({"accepted":False})
    else:
        return BAD_METHOD

@CheckRequire
def auto_validate(req: HttpRequest, id: any):
    if req.method == "PUT":
        body = json.loads(req.body.decode())
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        id = require({"id":id},"id","int","bad param [id]",-1)
        package = TaskPackage.objects.filter(id=id)
        if not package.exists():
            return request_failed(1,"Package not found",404)
        package=package.first()
        sender = User.objects.get(id=body['senderId'])
        if sender != package.creator:
            return request_failed(2,"No permission to validate package",401)
        if package.distributed == False:
            return request_failed(-2,"Cannot validate an undistributed package")
        tasks = Task.objects.filter(packageBelonging=package,hasStandardAnswer=True)
        if not tasks.exists():
            return request_failed(-2,"Cannot auto validate because no standard answers provided.")
        sender.lastupdatedtime = get_timestamp()
        sender.save()
        completedUser = package.completedUser
        validationRes=[]
        for user in completedUser.all():
            correctCount=0
            for task in tasks.all():
                answer = TaskAnswer.objects.filter(submitter=user,task=task)
                if not answer.exists():
                    return request_failed(1,"The user's answer not found. Contact admin about this.",500)
                answer=answer.first()
                correctCount += (answer.payload == task.standardAnswer)
            if correctCount >= tasks.count() * calAcceptRate(user):
                package.verifiedUser.add(user)
                result_bonus_calculate(user.id,id,True)
                validationRes.append({"userId":user.id,"accepted":True})
            else:
                result_bonus_calculate(user.id,id,False)
                package.availbleCount += 1
                validationRes.append({"userId":user.id,"accepted":False})
        for user in completedUser.all():
            package.completedUser.remove(user)
        package.save()
        return request_success({"result":validationRes})
    else:
        return BAD_METHOD

@CheckRequire
def get_completed_user_list(req: HttpRequest, id: any):
    if req.method == "GET":
        body = req.GET.dict()
        packageId = require({"id":id}, "id", "int", "bad param [id]",-1)
        package = TaskPackage.objects.filter(id = packageId).first()
        if package == None:
            return request_failed(1, "Package not found", 404)

        authRes = userAuth(body)
        if authRes != None:
            return authRes

        sender_id = require(body, "senderId", "int", err_msg="Missing or error type of [senderId]", err_code=-2)
        sender = User.objects.filter(id = sender_id).first()
        if sender != package.creator and not sender.adminPrivilege:
            return request_failed(2, "No permission to validate package", 401)
        sender.lastupdatedtime = get_timestamp()
        sender.save()

        return request_success({"userList": [user.serialize() for user in package.completedUser.all()]})
    else:
        return BAD_METHOD

@CheckRequire
def export_result(req: HttpRequest, id: any):
    if req.method == "GET":
        id = require({"id":id},"id","int","bad param [id]",-1)
        package = TaskPackage.objects.filter(id=id).first()
        if package == None:
            return request_failed(1,"Package not found",404)
        body = req.GET.dict()
        join = require(body,"join","int","bad param [join]")
        if join != 0 and join != 1:
            return request_failed(-2,"bad param [join] not 0 or 1")
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        sender = User.objects.get(id=body['senderId'])
        if sender != package.creator:
            return request_failed(2,"No privilege to export result",401)
        
        if join == 1:
            template = package.template
            for obj in template.objectList:
                if (not obj['type'] in SUPPORTED_CONTENTS_WITHOUT_OUTPUTFILE) and obj['type']!='singlechoice' and obj['type']!='multiplechoice':
                    return request_failed(-2,"Unsupported module "+obj['type']+' for answer join')
        if not package.distributed:
            return request_failed(-2,"Package not distributed. Please distribute it first.")
        if not package.verifiedUser.exists():
            return request_failed(-2,"No answers has been validated. Please do the validation first.")
        folder_path = os.path.join(settings.MEDIA_ROOT, str(id))
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        else:
            shutil.rmtree(folder_path)
            os.mkdir(folder_path)
        tasks = Task.objects.filter(packageBelonging = package)
        for task in tasks.iterator():
            if join == 0:
                answerOfTask = []
                for user in package.verifiedUser.iterator():
                    answer = TaskAnswer.objects.filter(task=task,submitter=user).first()
                    if answer == None:
                        shutil.rmtree(folder_path)
                        return request_failed(1,"Answer not found. Contact admin about this.",500)
                    answerOfTask.append(answer.payload)
            elif join == 1:
                answerOfTask = []
                choiceCount = []
                for i in range(len(template.objectList)):
                    obj=template.objectList[i]
                    if obj['type']=='singlechoice' or obj['type']=='multiplechoice':
                        choiceCount.append([0]*obj['count'])
                    else:
                        choiceCount.append([])
                for user in package.verifiedUser.iterator():
                    answer = TaskAnswer.objects.filter(task=task,submitter=user).first()
                    if answer == None:
                        shutil.rmtree(folder_path)
                        return request_failed(1,"Answer not found. Contact admin about this.",500)
                    for i in range(len(template.objectList)):
                        obj=template.objectList[i]
                        if obj['type']=='singlechoice' or obj['type']=='multiplechoice':
                            for j in range(obj['count']):
                                choiceCount[i][j]+=int(answer.payload[i][j]['data'])
                usercount = package.completedUser.count()
                for i in range(len(template.objectList)):
                    obj=template.objectList[i]
                    answerOfTask.append([])
                    if obj['type']=='singlechoice':
                        maxChoice = choiceCount[i].index(max(choiceCount[i]))
                        for j in range(obj['count']):
                            answerOfTask[i].append({'data':(maxChoice==j)})
                    elif obj['type']=='multiplechoice':
                        for j in range(obj['count']):
                            answerOfTask[i].append({'data':(choiceCount[i][j]*2>=usercount)})
            json.dump(answerOfTask,open(os.path.join(folder_path,str(task.problemId)+".json"),"w"))
        
        output_filename = "".join(random.choices(string.digits+string.ascii_letters,k=32))
        output_url = 'media/'+output_filename+".zip"
        shutil.make_archive(os.path.join(settings.MEDIA_ROOT,output_filename), 'zip', folder_path)
        shutil.rmtree(folder_path)
        delete_file_async(os.path.join(settings.MEDIA_ROOT,output_filename+".zip"))
        return request_success({"url":output_url})
    else:
        return BAD_METHOD
    
@CheckRequire
def post_excel(req: HttpRequest):
    if req.method == "POST":
        sender_id = req.POST.get("senderId")
        sender_token = req.POST.get("senderToken")
        sender = User.objects.filter(id=sender_id).first()
        if not sender:
            return request_failed(1,"User not found",404)
        if sender.token != sender_token:
            return request_failed(2,"Token not match",401)
        if not sender.labelPrivilege:
            return request_failed(2,"No privilege to post excel",401)
        if sender.currentTaskPackage == None:
            return request_failed(-2,"You haven't got a package. Maybe your task has passed deadline.")
        
        _file = req.FILES.get("uploadFile", None)
        if not _file:
            return request_failed(-2,"No file uploaded", status_code=400)
        
        file_path = os.path.join(settings.MEDIA_ROOT, str(sender_id) + "_" + str(get_timestamp()) + ".xlsx")
        upload_file = open(file_path, "wb")
        for chunk in _file.chunks():
            upload_file.write(chunk)
        upload_file.close()

        def request_failed_with_file_delete(code: int, message: str, status_code: int = 400):
            os.remove(file_path)
            return request_failed(code,message,status_code)
        
        try:
            wb = openpyxl.load_workbook(file_path)
        except:
            request_failed_with_file_delete(-2, "Bad excel file")

        sheet = wb.active

        # if len(answers) != sheet.max_row:
            # return request_failed_with_file_delete(-2,"Answer count doesn't match. You may have submitted problem you haven't viewed or submit less answers than viewed.")
        for (i, row) in enumerate(sheet.iter_rows()):
            row[0].value = require({'data':row[0].value},'data','int','Id is not a integer.')
            task = Task.objects.filter(packageBelonging=sender.currentTaskPackage,problemId=row[0].value).first()
            if task == None:
                return request_failed_with_file_delete(-2,f'The problem {row[0].value} does not exist.')
            answer = TaskAnswer.objects.filter(submitter=sender,task=task).first()
            if answer != None and answer.finished:
                return request_failed_with_file_delete(-2,f'Problem {row[0].value} is already completed')
            template = sender.currentTaskPackage.template
            now = 1
            for j in range(len(template.objectList)):
                unitType = template.objectList[j]['type']

                if unitType in SUPPORTED_CONTENTS_WITHOUT_OUTPUTFILE:
                    continue
                if not unitType in SUPPORTED_CONTENTS_WITHOUT_INPUTFILE:
                    return request_failed_with_file_delete(-2,"Unsupported content type")
                sub_payload = []
                count = 0
                for _ in range(template.objectList[j]['count']):
                    if now >= len(row):
                        return request_failed_with_file_delete(-2,"format of answer doesn't match with template [0x1]")
                    if unitType == "singlechoice":
                        if str(row[now].value) != "True" and str(row[now].value) != "False":
                            return request_failed_with_file_delete(-2,"format of answer doesn't match with template [0x2]")
                        count += int("True" == str(row[now].value))
                    elif unitType == "multiplechoice":
                        if str(row[now].value) != "True" and str(row[now].value) != "False":
                            return request_failed_with_file_delete(-2,"format of answer doesn't match with template [0x3]")
                    now += 1
                if unitType == "singlechoice" and count != 1:
                    return request_failed_with_file_delete(-2,"format of answer doesn't match with template [0x4]")
                if now != len(row):
                    return request_failed_with_file_delete(-2,"format of answer doesn't match with template [0x5]")

        for (i, row) in enumerate(sheet.iter_rows()):
            payload = []
            task = Task.objects.filter(packageBelonging=sender.currentTaskPackage,problemId=row[0].value).first()
            answer = TaskAnswer.objects.filter(submitter=sender,task=task).first()
            if answer == None:
                answer=TaskAnswer.objects.create(task=task,submitter=sender)
            template = sender.currentTaskPackage.template
            now = 1
            for j in range(len(template.objectList)):
                unitType = template.objectList[j]['type']
                if unitType in SUPPORTED_CONTENTS_WITHOUT_OUTPUTFILE:
                    payload.append([])
                    continue
                sub_payload = []
                for _ in range(template.objectList[j]['count']):
                    if unitType == "textinput":
                        sub_payload.append({'data':str(row[now].value)})
                    elif unitType == "singlechoice":
                        sub_payload.append({'data':"True" == str(row[now].value)})
                    elif unitType == "multiplechoice":
                        sub_payload.append({'data':"True" == str(row[now].value)})
                    now += 1
                
                payload.append(sub_payload)
            answer.payload=payload
            answer.finished=True
            answer.save()

        sender.lastupdatedtime = get_timestamp()
        sender.save()
        return request_success()
    else:
        return BAD_METHOD

@CheckRequire
def get_task_list(req: HttpRequest, id: any):
    if req.method == "GET":
        body = req.GET.dict()
        authRes = userAuth(body)
        if authRes != None:
            return authRes
        id=require({"id":id},"id","int","bad param [id]")
        
        pageId = require(body,'pageId','int',err_msg="bad param [pageId]")
        if pageId <= 0:
            return request_failed(-2,"bad param [pageId] with pageId leq to 0")
        count = require(body,'count','int',err_msg='bad param [cound]')
        if count <=0:
            return request_failed(-2,"bad param [count] with count leq to 0")
        if count > MAX_TASK_LIST_COUNT:
            return request_failed(-2,"bad param [count] with too large count")
        package=TaskPackage.objects.filter(id=id).first()
        if package==None:
            return request_failed(1,"Package not found!",404)
        sender = User.objects.get(id=body['senderId'])
        if sender.currentTaskPackage == package and sender.deadline < get_timestamp():
            handle_user_timeout(sender)
        if not sender.adminPrivilege and sender != package.creator and sender.currentTaskPackage != package:
            return request_failed(2,"No permission!",401)
        tasks = Task.objects.filter(packageBelonging=package).order_by("problemId")
        pages = Paginator(tasks,count)
        try:
            page = pages.page(pageId)
        except EmptyPage:
            page = pages.page(pages.num_pages)
        sender.refresh_from_db()
        sender.lastupdatedtime=get_timestamp()
        sender.save()
        def get_completed_state(task:Task):
            answer = TaskAnswer.objects.filter(submitter=sender,task=task).first()
            if answer==None:
                return 0
            if answer.finished:
                return 2
            # if answer.startTime + package.subtaskLimit < get_timestamp():
            #     return 1
            return 0
        return request_success({
            "taskList":[{
                "data": t.data,
                "problemId": t.problemId,
                "completed": get_completed_state(t),
            } for t in page.object_list],
            "pageCount":pages.num_pages,
            "template":package.template.serialize(),
            })
    else:
        return BAD_METHOD